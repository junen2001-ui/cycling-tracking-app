from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def _format_header(cell):
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')


def write_excel_report(output_path: Path, daily_rows: Iterable[dict[str, object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Daily Summary'

    headers = ['日付', '合計評価額', '増減', 'NISA', '投資信託', '個別株', 'その他']
    sheet.append(headers)
    for cell in sheet[1]:
        _format_header(cell)

    for row in daily_rows:
        sheet.append([
            row['date'],
            row['total_value'],
            row['change'],
            row['NISA'],
            row['投資信託'],
            row['個別株'],
            row['その他'],
        ])

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=len(headers), max_row=sheet.max_row):
        date_cell = row[0]
        date_cell.number_format = 'yyyy-mm-dd'
        for value_cell in row[1:]:
            value_cell.number_format = '#,##0.00'

    for idx, width in enumerate((12, 16, 12, 14, 14, 14, 12), start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    chart = LineChart()
    chart.title = '日次資産評価額推移'
    chart.style = 12
    chart.y_axis.title = '評価額'
    chart.x_axis.title = '日付'
    chart.x_axis.number_format = 'yyyy-mm-dd'
    chart.x_axis.majorTimeUnit = 'days'

    data = Reference(sheet, min_col=2, max_col=6, min_row=1, max_row=sheet.max_row)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 15
    chart.width = 32
    sheet.add_chart(chart, 'I2')

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
